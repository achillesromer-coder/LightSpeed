from __future__ import annotations

import csv
import hashlib
import json
import math
from datetime import datetime, timezone
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


OUT_ROOT = Path("romer_rfs_emff_type1_analysis_2026-08-05")
FIG_DIR = OUT_ROOT / "figures"
SOURCE_PDF = Path(r"C:\Users\acc\Desktop\2025 Documents\Romer Ind\Project Overview.pdf")
PRIOR_RECEIPT = Path("rfs_emff_sweep_proof_2026-08-05") / "sweep_receipt.json"

C0 = 299_792_458.0
MU0 = 4.0 * math.pi * 1e-7


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


materials = [
    {
        "material_id": "Au",
        "name": "Gold",
        "pdf_resonance_cm1_min": 1000,
        "pdf_resonance_cm1_max": 2000,
        "density_kg_m3": 19320,
        "conductivity_s_m": 4.10e7,
        "relative_permeability": 1.0,
        "susceptibility_proxy": -3.4e-5,
        "target_role": "primary noble/electric-field responsive target",
        "confidence": "reference-required",
        "source": "Project Overview p4; NIST resistivity literature baseline; engineering density default",
    },
    {
        "material_id": "Nd",
        "name": "Neodymium",
        "pdf_resonance_cm1_min": 400,
        "pdf_resonance_cm1_max": 4000,
        "density_kg_m3": 7007,
        "conductivity_s_m": 1.56e6,
        "relative_permeability": 1.05,
        "susceptibility_proxy": 5.6e-3,
        "target_role": "secondary rare-earth magnetic-response target",
        "confidence": "placeholder-needs-material-certificate",
        "source": "Project Overview p4; values must be replaced by assay/vendor certificate",
    },
    {
        "material_id": "Fe",
        "name": "Iron",
        "pdf_resonance_cm1_min": 200,
        "pdf_resonance_cm1_max": 700,
        "density_kg_m3": 7874,
        "conductivity_s_m": 1.00e7,
        "relative_permeability": 200.0,
        "susceptibility_proxy": 2.0e2,
        "target_role": "primary magnetic control/target",
        "confidence": "screening-default",
        "source": "Project Overview p4; NIST resistivity literature baseline; permeability highly state-dependent",
    },
    {
        "material_id": "C_graphite",
        "name": "Carbon / graphite",
        "pdf_resonance_cm1_min": 2800,
        "pdf_resonance_cm1_max": 3300,
        "density_kg_m3": 2260,
        "conductivity_s_m": 1.0e5,
        "relative_permeability": 1.0,
        "susceptibility_proxy": -2.0e-5,
        "target_role": "control and carbonaceous/regolith proxy",
        "confidence": "anisotropic-placeholder",
        "source": "Project Overview p4; graphite properties anisotropic; use measured sample values before claims",
    },
]

samples = [
    ("1a", "Au", "Gravel dust", "variable dust particle size", "find optimal wavelengths", 1600, 0.010),
    ("1b", "Au", "Clay sphere", "variable clay grit and moisture", "find permeability limitation", 1800, 0.010),
    ("1c", "Au", "Anthropic regolith", "variable regolith density", "find optimal wavelengths/displacement", 1500, 0.010),
    ("2a", "Nd", "Sands", "variable grain size / grit and purity", "find optimal SOP", 1650, 0.010),
    ("2b", "Nd", "Clay ball", "variable clay density, optimal grit", "find permeability limitation", 1800, 0.010),
    ("2c", "Nd", "Anthropic regolith", "variable element concentrations", "find range of operations", 1500, 0.010),
    ("3a", "Fe", "Sand", "variable element concentration", "find optimal displacement", 1650, 0.010),
    ("3b", "Fe", "Clay ball", "variable element concentration", "find optimal energy/displacement", 1800, 0.010),
    ("3c", "Fe", "Anthropic regolith", "statistically improbable natural samples", "find optimal SOP/energy/extraction", 1500, 0.010),
    ("4a", "C_graphite", "Graphene / gravel", "variable element concentration", "map energy pathways", 1550, 0.010),
    ("4b", "C_graphite", "Baked clay ball", "variable element purity and alloys", "find range of operations", 1900, 0.010),
    ("4c", "Mixed", "Anthropic regolith element mixes", "statistically most probable natural samples", "base training and deployment planning/SOP", 1500, 0.010),
]

field_t_values = [0.05, 0.10, 0.20, 0.50]
thickness_m_values = [0.001, 0.005, 0.010]
particle_diameter_m_values = [1e-5, 1e-4, 1e-3]
concentration_wt_values = [0.001, 0.010, 0.050, 0.100]
applied_wavenumber_cm1_values = [200, 400, 700, 1000, 1500, 2000, 2800, 3300, 4000]
rf_frequency_mhz_values = [1, 2, 5, 10, 20, 50]
base_volume_l = 1.0


def band_center(row: dict) -> float:
    return (float(row["pdf_resonance_cm1_min"]) + float(row["pdf_resonance_cm1_max"])) / 2.0


def resonance_alignment(applied_cm1: float, lo: float, hi: float) -> float:
    center = (lo + hi) / 2.0
    half_width = max((hi - lo) / 2.0, 1.0)
    distance = max(0.0, abs(applied_cm1 - center) - half_width)
    return math.exp(-distance / max(center * 0.15, 1.0))


def skin_depth_m(freq_hz: float, sigma: float, mu_r: float) -> float:
    omega = 2 * math.pi * max(freq_hz, 1e-9)
    return math.sqrt(2.0 / max(omega * MU0 * mu_r * sigma, 1e-30))


def emff_response(freq_hz: float, field_t: float, thickness_m: float, sigma: float, mu_r: float) -> tuple[float, float, float]:
    delta = skin_depth_m(freq_hz, sigma, mu_r)
    attenuation = math.exp(-thickness_m / max(delta, 1e-12))
    response = (field_t * attenuation) / max(math.log10(1 + freq_hz), 1e-12)
    return delta, attenuation, response


def build_frames() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    mat_df = pd.DataFrame(materials)
    sample_rows = []
    for sample_id, target, medium, variables, objective, bulk_density, volume_l in samples:
        sample_rows.append(
            {
                "sample_id": sample_id,
                "target_material_id": target,
                "medium": medium,
                "variable_focus": variables,
                "objective": objective,
                "base_volume_l": volume_l * 1000 if volume_l < 1 else volume_l,
                "base_volume_m3": base_volume_l / 1000.0,
                "bulk_density_kg_m3": bulk_density,
                "bulk_mass_kg": bulk_density * (base_volume_l / 1000.0),
                "source": "Project Overview p5; bulk density/default volume are screening assumptions",
            }
        )
    sample_df = pd.DataFrame(sample_rows)

    sweep_rows = []
    mat_lookup = {m["material_id"]: m for m in materials}
    for _, sample in sample_df.iterrows():
        candidates = materials if sample["target_material_id"] == "Mixed" else [mat_lookup[sample["target_material_id"]]]
        for mat in candidates:
            for applied_cm1 in applied_wavenumber_cm1_values:
                align = resonance_alignment(applied_cm1, mat["pdf_resonance_cm1_min"], mat["pdf_resonance_cm1_max"])
                optical_freq_hz = applied_cm1 * 100.0 * C0
                for rf_mhz in rf_frequency_mhz_values:
                    rf_hz = rf_mhz * 1e6
                    rf_cm1_equiv = rf_hz / (100.0 * C0)
                    rf_to_band_ratio = rf_cm1_equiv / max(band_center(mat), 1e-9)
                    for field_t in field_t_values:
                        for thickness in thickness_m_values:
                            delta, atten, resp = emff_response(
                                rf_hz,
                                field_t,
                                thickness,
                                mat["conductivity_s_m"],
                                mat["relative_permeability"],
                            )
                            for diameter in particle_diameter_m_values:
                                particle_volume = 4 / 3 * math.pi * (diameter / 2) ** 3
                                particle_mass = particle_volume * mat["density_kg_m3"]
                                for concentration in concentration_wt_values:
                                    target_mass = sample["bulk_mass_kg"] * concentration
                                    mass_proxy = min(1.0, max(0.0, concentration / 0.05))
                                    magnetic_proxy = min(1.0, abs(mat["susceptibility_proxy"]) / 10.0)
                                    screening_score = align * (field_t / max(field_t_values)) * mass_proxy * (0.35 + 0.65 * magnetic_proxy)
                                    penetration_score = min(1.0, delta / max(thickness, 1e-12))
                                    combined_score = screening_score * (0.2 + 0.8 * penetration_score)
                                    sweep_rows.append(
                                        {
                                            "sample_id": sample["sample_id"],
                                            "material_id": mat["material_id"],
                                            "applied_wavenumber_cm1": applied_cm1,
                                            "band_center_cm1": band_center(mat),
                                            "resonance_alignment_0_1": align,
                                            "optical_equiv_frequency_hz": optical_freq_hz,
                                            "rf_frequency_mhz": rf_mhz,
                                            "rf_equiv_wavenumber_cm1": rf_cm1_equiv,
                                            "rf_to_pdf_band_ratio": rf_to_band_ratio,
                                            "field_t": field_t,
                                            "thickness_m": thickness,
                                            "skin_depth_m": delta,
                                            "attenuation_0_1": atten,
                                            "ls_runner_response_proxy": resp,
                                            "particle_diameter_m": diameter,
                                            "particle_mass_kg": particle_mass,
                                            "concentration_wt": concentration,
                                            "target_mass_kg": target_mass,
                                            "screening_score_0_1": screening_score,
                                            "combined_rank_score_0_1": combined_score,
                                            "claim_gate": "screening_only",
                                        }
                                    )
    sweep_df = pd.DataFrame(sweep_rows)
    best_df = (
        sweep_df.sort_values(["combined_rank_score_0_1", "resonance_alignment_0_1"], ascending=False)
        .groupby(["sample_id", "material_id"], as_index=False)
        .head(1)
        .reset_index(drop=True)
    )
    return mat_df, sample_df, sweep_df, best_df


def write_csvs(mat_df: pd.DataFrame, sample_df: pd.DataFrame, sweep_df: pd.DataFrame, best_df: pd.DataFrame) -> None:
    mat_df.to_csv(OUT_ROOT / "material_properties.csv", index=False)
    sample_df.to_csv(OUT_ROOT / "sample_matrix.csv", index=False)
    sweep_df.to_csv(OUT_ROOT / "simulation_grid.csv", index=False)
    best_df.to_csv(OUT_ROOT / "best_screening_cases.csv", index=False)


def make_figures(sweep_df: pd.DataFrame, best_df: pd.DataFrame) -> list[Path]:
    paths: list[Path] = []
    top = best_df.sort_values("combined_rank_score_0_1", ascending=False).head(16)
    labels = top["sample_id"] + "-" + top["material_id"]
    plt.figure(figsize=(12, 6))
    plt.bar(labels, top["combined_rank_score_0_1"], color="#0f766e")
    plt.xticks(rotation=45, ha="right")
    plt.ylabel("Combined screening score (0-1)")
    plt.title("Top bounded RFS/EMFF screening cases")
    plt.tight_layout()
    p = FIG_DIR / "top_screening_cases.png"
    plt.savefig(p, dpi=180)
    plt.close()
    paths.append(p)

    pivot = (
        sweep_df.groupby(["material_id", "applied_wavenumber_cm1"])["resonance_alignment_0_1"]
        .max()
        .unstack(fill_value=0)
        .reindex([m["material_id"] for m in materials])
    )
    plt.figure(figsize=(12, 4.5))
    plt.imshow(pivot.values, aspect="auto", cmap="viridis", vmin=0, vmax=1)
    plt.colorbar(label="Resonance alignment (0-1)")
    plt.yticks(range(len(pivot.index)), pivot.index)
    plt.xticks(range(len(pivot.columns)), [str(int(c)) for c in pivot.columns], rotation=45)
    plt.xlabel("Applied wavenumber (cm^-1)")
    plt.title("PDF resonance-band alignment sweep")
    plt.tight_layout()
    p = FIG_DIR / "resonance_alignment_heatmap.png"
    plt.savefig(p, dpi=180)
    plt.close()
    paths.append(p)

    skin = sweep_df[(sweep_df["thickness_m"] == 0.001) & (sweep_df["field_t"] == 0.10)]
    skin_p = skin.groupby(["material_id", "rf_frequency_mhz"])["skin_depth_m"].median().unstack()
    plt.figure(figsize=(10, 5))
    for material_id in skin_p.index:
        plt.plot(skin_p.columns, skin_p.loc[material_id], marker="o", label=material_id)
    plt.yscale("log")
    plt.xlabel("RF/coil frequency (MHz)")
    plt.ylabel("Skin depth (m, log scale)")
    plt.title("EMFF/RF penetration proxy by material")
    plt.legend()
    plt.tight_layout()
    p = FIG_DIR / "skin_depth_by_material.png"
    plt.savefig(p, dpi=180)
    plt.close()
    paths.append(p)

    mismatch = best_df.copy()
    mismatch["rf_ratio_log10"] = np.log10(np.maximum(mismatch["rf_to_pdf_band_ratio"], 1e-20))
    plt.figure(figsize=(10, 5))
    plt.scatter(mismatch["band_center_cm1"], mismatch["rf_ratio_log10"], c=mismatch["combined_rank_score_0_1"], cmap="plasma")
    plt.colorbar(label="Combined score")
    plt.xlabel("PDF resonance band center (cm^-1)")
    plt.ylabel("log10(RF equivalent wavenumber / band center)")
    plt.title("MHz RF domain is many orders below PDF cm^-1 bands")
    plt.tight_layout()
    p = FIG_DIR / "rf_vs_pdf_frequency_domain_gap.png"
    plt.savefig(p, dpi=180)
    plt.close()
    paths.append(p)

    return paths


def style_ws(ws) -> None:
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E4E2")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = Border(bottom=thin)
    ws.freeze_panes = "A2"
    for col in range(1, min(ws.max_column, 18) + 1):
        values = [ws.cell(row=r, column=col).value for r in range(1, min(ws.max_row, 80) + 1)]
        width = min(max(len(str(v)) if v is not None else 0 for v in values) + 2, 42)
        ws.column_dimensions[get_column_letter(col)].width = max(width, 10)


def add_df_sheet(wb: Workbook, title: str, df: pd.DataFrame, max_rows: int | None = None) -> None:
    ws = wb.create_sheet(title)
    data = df if max_rows is None else df.head(max_rows)
    ws.append(list(data.columns))
    for row in data.itertuples(index=False):
        ws.append(list(row))
    style_ws(ws)
    if ws.max_row > 1 and ws.max_column > 1:
        table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName=title.replace(" ", "_").replace("-", "_")[:25], ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)


def build_workbook(mat_df: pd.DataFrame, sample_df: pd.DataFrame, sweep_df: pd.DataFrame, best_df: pd.DataFrame, fig_paths: list[Path]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws.append(["Field", "Value"])
    rows = [
        ("Purpose", "Type-1 Cognigrex-compatible RFS/EMFF screening workbook for Romer/ECO-X/EMASSC sample planning."),
        ("Scope", "Pre-alloy, pre-concentration-refinement starting matrix from Project Overview plus bounded local simulation."),
        ("Claim boundary", "Screening/proofing only; no extraction-rate, purity, optimal-frequency, energy-efficiency, or equipment-safety claim."),
        ("Source PDF", str(SOURCE_PDF)),
        ("PDF status", "Image-based PDF; pages rendered locally; values transcribed from visible slides."),
        ("Prior LS proof receipt", str(PRIOR_RECEIPT)),
        ("Generated UTC", datetime.now(timezone.utc).isoformat()),
        ("Core shorthand equation", "Score = evidence_gate * sample_gate * resonance_alignment * field_factor * concentration_factor * penetration_factor"),
        ("Resonance alignment", "exp(-max(0, abs(applied_cm1 - band_center)-band_half_width)/(0.15*band_center))"),
        ("Penetration factor", "min(1, skin_depth / sample_thickness), skin_depth=sqrt(2/(2*pi*f*mu*sigma))"),
        ("Important gap", "PDF cm^-1 resonance bands correspond to THz optical/IR frequencies; existing LS MHz coil sweep is a separate RF/EMFF penetration domain."),
    ]
    for r in rows:
        ws.append(list(r))
    style_ws(ws)

    add_df_sheet(wb, "Materials", mat_df)
    add_df_sheet(wb, "Samples", sample_df)
    add_df_sheet(wb, "Best Cases", best_df)
    add_df_sheet(wb, "Sweep Grid", sweep_df, max_rows=5000)

    ws_calc = wb.create_sheet("Formula Sandbox")
    sandbox_rows = [
        ["Input", "Value", "Formula/Note"],
        ["sample_volume_m3", base_volume_l / 1000, "Editable base volume"],
        ["bulk_density_kg_m3", 1500, "Editable medium bulk density"],
        ["target_concentration_wt", 0.01, "Editable target concentration"],
        ["sample_mass_kg", "=B2*B3", "base volume*density"],
        ["target_mass_kg", "=B5*B4", "sample mass*concentration"],
        ["frequency_hz", 1000000, "Editable RF/coil frequency"],
        ["conductivity_s_m", 41000000, "Editable material conductivity"],
        ["relative_permeability", 1, "Editable material mu_r"],
        ["sample_thickness_m", 0.001, "Editable sample bed thickness"],
        ["skin_depth_m", "=SQRT(2/(2*PI()*B7*4*PI()*1E-7*B9*B8))", "RF/EMFF penetration proxy"],
        ["penetration_factor", "=MIN(1,B11/B10)", "bounded 0-1"],
    ]
    for row in sandbox_rows:
        ws_calc.append(row)
    style_ws(ws_calc)

    ws_results = wb.create_sheet("Results Dashboard")
    ws_results.append(["Metric", "Value", "Interpretation"])
    dashboard_rows = [
        ["Material count", len(mat_df), "PDF starting set before alloy expansion"],
        ["Sample count", len(sample_df), "Visible sample IDs from Project Overview"],
        ["Sweep rows generated", len(sweep_df), "Local bounded simulation grid"],
        ["Best case max score", float(best_df["combined_rank_score_0_1"].max()), "Relative screening score, not extraction efficiency"],
        ["Physical claim allowed", 0, "No empirical replicated raw logs attached in this workbook"],
        ["Canonical readiness", "draft-review", "Ready for LS review/control, not public canonical claim"],
    ]
    for r in dashboard_rows:
        ws_results.append(r)
    style_ws(ws_results)

    best_start = 10
    ws_results.cell(best_start, 1, "Top cases")
    for c, name in enumerate(["Sample", "Material", "Applied cm^-1", "Field T", "RF MHz", "Score"], 1):
        ws_results.cell(best_start + 1, c, name)
    for r_i, row in enumerate(best_df.sort_values("combined_rank_score_0_1", ascending=False).head(12).itertuples(index=False), best_start + 2):
        ws_results.cell(r_i, 1, row.sample_id)
        ws_results.cell(r_i, 2, row.material_id)
        ws_results.cell(r_i, 3, row.applied_wavenumber_cm1)
        ws_results.cell(r_i, 4, row.field_t)
        ws_results.cell(r_i, 5, row.rf_frequency_mhz)
        ws_results.cell(r_i, 6, row.combined_rank_score_0_1)
    style_ws(ws_results)

    chart = BarChart()
    chart.title = "Top screening scores"
    chart.y_axis.title = "Score"
    chart.x_axis.title = "Case"
    data = Reference(ws_results, min_col=6, min_row=best_start + 1, max_row=best_start + 13)
    cats = Reference(ws_results, min_col=1, min_row=best_start + 2, max_row=best_start + 13)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    ws_results.add_chart(chart, "H10")

    ws_sources = wb.create_sheet("Sources")
    ws_sources.append(["Source", "URL/Path", "Use"])
    for row in [
        ["Project Overview.pdf", str(SOURCE_PDF), "RFS/EMFF concept, sample elements, sample IDs, test conditions"],
        ["NASA Psyche", "https://science.nasa.gov/solar-system/asteroids/16-psyche/", "Metal/silicate range for metallic asteroid target framing"],
        ["NASA Bennu sample", "https://www.nasa.gov/news-release/nasas-bennu-asteroid-sample-contains-carbon-water/", "Carbon/water-rich returned sample context"],
        ["NIST resistivity", "https://srd.nist.gov/jpcrdreprint/1.555614.pdf", "Electrical resistivity baseline literature"],
        ["Prior LS receipt", str(PRIOR_RECEIPT), "Evidence gate and local RFS/EMFF sweep proof"],
    ]:
        ws_sources.append(row)
    style_ws(ws_sources)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    out_path = OUT_ROOT / "Romer_RFS_EMFF_Type1_Cognigrex_Workbook_2026-08-05.xlsx"
    wb.save(out_path)
    return out_path


def write_handoff(mat_df: pd.DataFrame, sample_df: pd.DataFrame, sweep_df: pd.DataFrame, best_df: pd.DataFrame, workbook_path: Path, fig_paths: list[Path]) -> Path:
    handoff = OUT_ROOT / "RFS_EMFF_TYPE1_HANDOFF_AND_WORKFLOW_2026-08-05.md"
    top = best_df.sort_values("combined_rank_score_0_1", ascending=False).head(8)
    lines = [
        "# RFS/EMFF Type-1 Handoff and Workflow - 2026-08-05",
        "",
        "## Execution boundary",
        "",
        "This package is a bounded screening and planning artifact. It does not claim extraction rate, purity, optimal frequency, energy efficiency, equipment safety, or validated physical capability.",
        "",
        "## Current outputs",
        "",
        f"- Workbook: `{workbook_path.resolve()}`",
        f"- Sweep rows: `{len(sweep_df)}`",
        f"- Materials: `{len(mat_df)}`",
        f"- Samples: `{len(sample_df)}`",
        f"- Figures: `{len(fig_paths)}`",
        "",
        "## Shorthand equation",
        "",
        "`Score = evidence_gate * sample_gate * resonance_alignment * field_factor * concentration_factor * penetration_factor`",
        "",
        "Where `penetration_factor = min(1, sqrt(2/(2*pi*f*mu*sigma)) / thickness)` and `resonance_alignment` is a bounded distance from the Project Overview cm^-1 resonance band.",
        "",
        "## Critical finding",
        "",
        "The Project Overview sample resonance bands are expressed in cm^-1, which correspond to THz optical/IR-scale frequencies. The existing LS runner uses MHz RF/coil sweeps for skin-depth and attenuation. These must remain separate model domains until the applied excitation modality is specified and instrumented.",
        "",
        "## Top screening cases",
        "",
        "| sample | material | applied cm^-1 | field T | rf MHz | score |",
        "|---|---:|---:|---:|---:|---:|",
    ]
    for row in top.itertuples(index=False):
        lines.append(
            f"| {row.sample_id} | {row.material_id} | {row.applied_wavenumber_cm1:g} | {row.field_t:g} | {row.rf_frequency_mhz:g} | {row.combined_rank_score_0_1:.4f} |"
        )
    lines.extend(
        [
            "",
            "## LS execution workflow",
            "",
            "1. Confirm base sample volume, bulk density, moisture, particle-size distribution, and target concentration for each sample ID.",
            "2. Replace placeholder material properties with measured assay/certificate values.",
            "3. Run RFS-domain resonance sweeps in cm^-1/THz-equivalent terms only when the actual excitation modality is defined.",
            "4. Run EMFF-domain MHz/coil sweeps using LS `rfs_emff_runner.py` for skin depth and field penetration comparisons.",
            "5. Attach instrumented repeated raw logs before allowing physical capability language.",
            "6. Promote to canonical LS/web only after Merovingian/Smith gates confirm evidence, artifact hashes, and claim boundaries.",
            "",
            "## Z-floor responsibilities",
            "",
            "- Z-3 Smith: validate claim gates, math domain, and blocked claims.",
            "- Z-4 Merovingian: register receipts/manifests and enforce bounded work intake.",
            "- Z+3 Trinity: maintain workflow/schema compatibility and operator handoff state.",
            "- Z0 TheConstruct: run bounded simulation jobs and export artifacts.",
            "",
            "## Known pinch points",
            "",
            "- PDF OCR is not embedded; values were transcribed from rendered pages.",
            "- All alloy/concentration variations remain parametric until sample assays are known.",
            "- Permeability and susceptibility are strongly material-state dependent, especially iron and neodymium.",
            "- Safety and equipment performance require physical instrumentation, not spreadsheet inference.",
        ]
    )
    handoff.write_text("\n".join(lines), encoding="utf-8")
    return handoff


def write_manifest(paths: dict[str, Path], data_summary: dict) -> Path:
    manifest_path = OUT_ROOT / "manifest.json"
    payload = {
        "schema": "romer-rfs-emff-type1-analysis-v1",
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "source_pdf": str(SOURCE_PDF),
        "prior_lightspeed_receipt": str(PRIOR_RECEIPT),
        "claim_boundary": "screening_only_no_physical_capability_claim",
        "summary": data_summary,
        "artifacts": {k: {"path": str(v.resolve()), "sha256": sha256(v), "size_bytes": v.stat().st_size} for k, v in paths.items()},
    }
    manifest_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return manifest_path


def main() -> None:
    OUT_ROOT.mkdir(exist_ok=True)
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    mat_df, sample_df, sweep_df, best_df = build_frames()
    write_csvs(mat_df, sample_df, sweep_df, best_df)
    fig_paths = make_figures(sweep_df, best_df)
    workbook_path = build_workbook(mat_df, sample_df, sweep_df, best_df, fig_paths)
    handoff_path = write_handoff(mat_df, sample_df, sweep_df, best_df, workbook_path, fig_paths)
    paths = {
        "workbook": workbook_path,
        "handoff": handoff_path,
        "material_properties_csv": OUT_ROOT / "material_properties.csv",
        "sample_matrix_csv": OUT_ROOT / "sample_matrix.csv",
        "simulation_grid_csv": OUT_ROOT / "simulation_grid.csv",
        "best_screening_cases_csv": OUT_ROOT / "best_screening_cases.csv",
    }
    for idx, fig in enumerate(fig_paths, 1):
        paths[f"figure_{idx}"] = fig
    manifest_path = write_manifest(
        paths,
        {
            "materials": len(mat_df),
            "samples": len(sample_df),
            "sweep_rows": len(sweep_df),
            "best_score": float(best_df["combined_rank_score_0_1"].max()),
            "physical_claim_allowed": False,
        },
    )
    print(json.dumps({"out_root": str(OUT_ROOT.resolve()), "workbook": str(workbook_path.resolve()), "handoff": str(handoff_path.resolve()), "manifest": str(manifest_path.resolve()), "rows": len(sweep_df)}, indent=2))


if __name__ == "__main__":
    main()
