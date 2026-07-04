from __future__ import annotations

from datetime import datetime
import json

from openpyxl import load_workbook
import pytest

from lightspeed_runtime.governance_ledgers import append_approval_ledger
from lightspeed_runtime.log_ledger_export import export_ledgers
from lightspeed_runtime.maintenance import (
    MaintenanceWindowClosed,
    run_maintenance,
)
from lightspeed_runtime.operational_store import (
    OperationalStore,
    default_operational_db_path,
)
from lightspeed_runtime.weekly_log import append_weekly_log


def test_maintenance_requires_friday_1900_without_force(tmp_path):
    with pytest.raises(MaintenanceWindowClosed):
        run_maintenance(
            tmp_path,
            now=datetime(2026, 7, 9, 19, 0),
        )


def test_maintenance_quarantines_generated_duplicates_only(tmp_path):
    source = tmp_path / "Data" / "empirical.fits"
    generated = tmp_path / "reports" / "run_20260701.json"
    source.parent.mkdir(parents=True)
    generated.parent.mkdir(parents=True)
    source.write_bytes(b"source")
    generated.write_text("{}\n", encoding="utf-8")

    result = run_maintenance(
        tmp_path,
        now=datetime(2026, 7, 10, 19, 0),
    )

    assert source.exists()
    assert not generated.exists()
    assert result["quarantined"] == 1
    assert result["manifest_path"].exists()
    manifest = json.loads(result["manifest_path"].read_text(encoding="utf-8"))
    assert manifest["entries"][0]["source_relative"] == "reports/run_20260701.json"
    assert manifest["entries"][0]["sha256"]
    assert manifest["entries"][0]["second_copy_verified"] is False


def test_maintenance_dry_run_writes_no_manifest_or_moves(tmp_path):
    generated = tmp_path / "reports" / "run_20260701.json"
    generated.parent.mkdir(parents=True)
    generated.write_text("{}\n", encoding="utf-8")

    result = run_maintenance(
        tmp_path,
        now=datetime(2026, 7, 10, 19, 0),
        dry_run=True,
    )

    assert generated.exists()
    assert result["quarantined"] == 0
    assert result["candidate_count"] == 1
    assert result["manifest_path"] is None


def test_operational_workbook_is_projection_from_sqlite_and_weekly_log(tmp_path):
    root = tmp_path / "LightSpeed"
    store = OperationalStore(default_operational_db_path(root))
    store.record_event(
        {
            "event_id": "task-1",
            "kind": "task",
            "source": "Neo",
            "target": "Smith",
            "status": "queued",
            "payload": {"title": "Review compact persistence"},
        }
    )
    append_approval_ledger(
        root,
        {
            "event_id": "approval-1",
            "source_floor": "Architect",
            "target_floor": "Neo",
            "action_id": "ACT-1",
            "approval_state": "approved",
        },
    )
    append_weekly_log(
        root,
        category="publish_review",
        message="Review public projection.",
        source="Neo",
        timestamp=datetime(2026, 7, 10, 19, 0),
    )
    output = tmp_path / "LightSpeed_Operational_Log_Ledger.xlsx"

    result = export_ledgers(root, output)

    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == [
        "Overview",
        "Tasks",
        "Handoffs",
        "Approvals",
        "Receipts",
        "Exceptions",
        "Daily Log",
    ]
    assert workbook["Tasks"].max_row == 2
    assert workbook["Approvals"].max_row == 2
    assert workbook["Daily Log"].max_row == 2
    assert result["operational_events"] == 2
