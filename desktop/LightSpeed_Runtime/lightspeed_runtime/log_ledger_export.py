from __future__ import annotations

import argparse
from datetime import UTC, datetime
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lightspeed_runtime.operational_store import (
    OperationalStore,
    default_operational_db_path,
)
from lightspeed_runtime.storage_paths import logs_root


CONSOLIDATION_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_ROOT = CONSOLIDATION_ROOT / "Desktop_Hooks" / "LightSpeed"
DEFAULT_OUTPUT = (
    CONSOLIDATION_ROOT
    / "Logs"
    / "Canonical"
    / "LightSpeed_Operational_Log_Ledger.xlsx"
)
SHEET_NAMES = (
    "Overview",
    "Tasks",
    "Handoffs",
    "Approvals",
    "Receipts",
    "Exceptions",
    "Daily Log",
)
EVENT_HEADERS = (
    "Event ID",
    "Kind",
    "Timestamp UTC",
    "Source",
    "Target",
    "Project",
    "Priority",
    "Risk",
    "Status",
    "Summary",
    "Payload JSON",
)
TEAL = "156162"
SECONDARY = "235160"
CHARCOAL = "0A0C0B"
CREAM = "F7F3E8"
GOLD = "C9A24A"


def _decoded_payload(row: dict[str, Any]) -> dict[str, Any]:
    try:
        stored = json.loads(str(row.get("payload_json") or "{}"))
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    payload = stored.get("payload")
    return payload if isinstance(payload, dict) else stored


def _summary(payload: dict[str, Any]) -> str:
    nested = payload.get("payload")
    nested = nested if isinstance(nested, dict) else {}
    status = payload.get("status")
    status = status if isinstance(status, dict) else {}
    for value in (
        payload.get("message"),
        payload.get("summary"),
        payload.get("title"),
        nested.get("title"),
        nested.get("message"),
        status.get("message"),
        payload.get("action_id"),
        payload.get("event"),
        payload.get("name"),
    ):
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def _event_row(row: dict[str, Any], payload: dict[str, Any]) -> list[Any]:
    return [
        row.get("event_id"),
        row.get("kind"),
        row.get("created_utc"),
        row.get("source"),
        row.get("target"),
        row.get("project_id"),
        row.get("priority"),
        row.get("risk"),
        row.get("status"),
        _summary(payload),
        json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str),
    ]


def _payload_signal(payload: dict[str, Any]) -> str:
    nested = payload.get("payload")
    nested = nested if isinstance(nested, dict) else {}
    return " ".join(
        str(value or "").lower()
        for value in (
            payload.get("kind"),
            payload.get("type"),
            payload.get("event"),
            payload.get("category"),
            nested.get("kind"),
            nested.get("type"),
        )
    )


def _is_task(row: dict[str, Any], payload: dict[str, Any]) -> bool:
    return row.get("kind") in {"task", "queue"} or "task" in _payload_signal(payload)


def _is_handoff(row: dict[str, Any], payload: dict[str, Any]) -> bool:
    return "handoff" in str(row.get("kind") or "").lower() or "handoff" in _payload_signal(payload)


def _is_approval(row: dict[str, Any], payload: dict[str, Any]) -> bool:
    return row.get("kind") == "governance_approval" or "approval" in _payload_signal(payload)


def _is_receipt(row: dict[str, Any], payload: dict[str, Any]) -> bool:
    signal = f"{row.get('kind', '')} {_payload_signal(payload)}".lower()
    return "receipt" in signal


def _is_exception(row: dict[str, Any], payload: dict[str, Any]) -> bool:
    signal = " ".join(
        str(value or "").lower()
        for value in (
            row.get("status"),
            row.get("risk"),
            payload.get("level"),
            payload.get("category"),
            _summary(payload),
        )
    )
    return any(marker in signal for marker in ("blocked", "critical", "error", "fail", "rejected"))


def _style_table(sheet, widths: Iterable[int]) -> None:
    header_fill = PatternFill("solid", fgColor=SECONDARY)
    for cell in sheet[1]:
        cell.font = Font(name="Garamond", size=11, color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Garamond", size=10, color=CHARCOAL)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False


def _append_event_sheet(sheet, records: list[tuple[dict[str, Any], dict[str, Any]]]) -> None:
    sheet.append(EVENT_HEADERS)
    for row, payload in records:
        sheet.append(_event_row(row, payload))
    _style_table(sheet, (30, 23, 27, 18, 18, 18, 12, 13, 14, 54, 100))


def _weekly_records(root: Path) -> tuple[list[Path], list[dict[str, Any]]]:
    paths = sorted(logs_root(root).glob("lightspeed_*_W*.jsonl"))
    records: list[dict[str, Any]] = []
    for path in paths:
        with path.open("r", encoding="utf-8", errors="replace") as handle:
            for line_number, raw in enumerate(handle, start=1):
                raw = raw.strip()
                if not raw:
                    continue
                try:
                    payload = json.loads(raw)
                except json.JSONDecodeError:
                    payload = {
                        "category": "unparsed",
                        "message": raw,
                        "level": "warning",
                    }
                payload["_source_file"] = str(path)
                payload["_source_line"] = line_number
                records.append(payload)
    return paths, records


def export_ledgers(runtime_root: Path, output_path: Path) -> dict[str, Any]:
    runtime_root = Path(runtime_root)
    database_path = default_operational_db_path(runtime_root)
    store = OperationalStore(database_path)
    rows = list(reversed(store.recent(limit=10_000)))
    decoded = [(row, _decoded_payload(row)) for row in rows]
    weekly_paths, weekly = _weekly_records(runtime_root)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    for name in SHEET_NAMES[1:]:
        workbook.create_sheet(name)

    categories = {
        "Tasks": [(row, payload) for row, payload in decoded if _is_task(row, payload)],
        "Handoffs": [(row, payload) for row, payload in decoded if _is_handoff(row, payload)],
        "Approvals": [(row, payload) for row, payload in decoded if _is_approval(row, payload)],
        "Receipts": [(row, payload) for row, payload in decoded if _is_receipt(row, payload)],
        "Exceptions": [(row, payload) for row, payload in decoded if _is_exception(row, payload)],
    }
    overview.append(["LightSpeed Operational Log Ledger", "Value"])
    overview_rows = (
        ("Generated UTC", datetime.now(UTC).isoformat()),
        ("Runtime root", str(runtime_root)),
        ("SQLite authority", str(database_path)),
        ("Operational events", len(rows)),
        ("Weekly ledger files", len(weekly_paths)),
        ("Weekly ledger rows", len(weekly)),
        ("Authority", "Projection only; SQLite and governed weekly JSONL remain authoritative."),
    )
    for record in overview_rows:
        overview.append(record)
    _style_table(overview, (34, 110))
    overview["A1"].fill = PatternFill("solid", fgColor=TEAL)
    overview["B1"].fill = PatternFill("solid", fgColor=TEAL)
    overview["A1"].font = Font(name="Garamond", size=14, color="FFFFFF", bold=True)
    overview["B1"].font = Font(name="Garamond", size=14, color=GOLD, bold=True)

    for name, records in categories.items():
        _append_event_sheet(workbook[name], records)

    daily = workbook["Daily Log"]
    daily.append(
        [
            "Timestamp",
            "Level",
            "Category",
            "Source",
            "Message",
            "Source File",
            "Source Line",
            "Payload JSON",
        ]
    )
    for payload in weekly:
        daily.append(
            [
                payload.get("timestamp"),
                payload.get("level"),
                payload.get("category"),
                payload.get("source"),
                payload.get("message"),
                payload.get("_source_file"),
                payload.get("_source_line"),
                json.dumps(payload.get("payload") or {}, ensure_ascii=False, sort_keys=True, default=str),
            ]
        )
    _style_table(daily, (27, 12, 25, 20, 60, 70, 12, 100))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {
        "output_path": str(output_path),
        "database_path": str(database_path),
        "operational_events": len(rows),
        "source_ledgers": len(weekly_paths),
        "rows": len(rows) + len(weekly),
        "sheet_rows": {
            name: workbook[name].max_row - 1
            for name in SHEET_NAMES
        },
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Export LightSpeed operational state to a reviewed workbook."
    )
    parser.add_argument("--runtime-root", type=Path, default=DEFAULT_ROOT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args(argv)
    print(
        json.dumps(
            export_ledgers(args.runtime_root.resolve(), args.output.resolve()),
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
