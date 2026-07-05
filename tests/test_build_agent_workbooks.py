from __future__ import annotations

import json
from pathlib import Path
import sys

from openpyxl import load_workbook
import pytest


REPO_ROOT = Path(__file__).resolve().parents[1]
TOOLS_ROOT = REPO_ROOT / "tools"
if str(TOOLS_ROOT) not in sys.path:
    sys.path.insert(0, str(TOOLS_ROOT))

from build_agent_workbooks import (
    FLOOR_SHEETS,
    NEO_SHEETS,
    build_floor_workbook,
    build_neo_workbook,
    validate_records,
    write_drive_artifacts,
)


def test_neo_workbook_has_approved_tabs(tmp_path):
    path = build_neo_workbook(tmp_path / "N_LightSpeed.xlsx", records=[])
    workbook = load_workbook(path, data_only=False)

    assert workbook.sheetnames == NEO_SHEETS
    assert workbook["05 Open Tasks"].freeze_panes == "A5"
    assert workbook["01 Index"]["A5"].hyperlink is not None


def test_floor_workbook_uses_raphael_classification_palette(tmp_path):
    path = build_floor_workbook(
        tmp_path / "Oracle_LightSpeed.xlsx",
        floor="Oracle",
        records=[],
    )
    workbook = load_workbook(path)
    legend = workbook["12 Legend"]
    colors = {
        legend.cell(row=row, column=2).fill.fgColor.rgb
        for row in range(5, 10)
    }

    assert workbook.sheetnames == FLOOR_SHEETS
    assert {
        "FFB7FF59",
        "FF3FA65B",
        "FFF2C94C",
        "FF5A2A4A",
        "FFD66A6A",
    } <= colors


def test_floor_workbook_preserves_flexible_extension_fields(tmp_path):
    record = {
        "Item_ID": "ORACLE-K-001",
        "Workstream": "Knowns",
        "Artifact": "Empirical register",
        "Owner": "Oracle",
        "Release_Class": "RI-INTERNAL",
        "Claim_Grade": "Known",
        "Status": "active",
        "Notes": "Reviewed",
        "Extensions": {"bell_curve": "overlap-17", "confidence": 0.94},
    }

    path = build_floor_workbook(
        tmp_path / "Oracle_LightSpeed.xlsx",
        floor="Oracle",
        records=[record],
    )
    workbook = load_workbook(path, data_only=False)
    sheet = workbook["04 Knowns"]
    headers = [cell.value for cell in sheet[4]]
    extension_column = headers.index("Extensions_JSON") + 1

    assert json.loads(sheet.cell(5, extension_column).value)["confidence"] == 0.94
    assert sheet.auto_filter.ref
    assert "A4" in sheet.auto_filter.ref


def test_duplicate_fact_ownership_is_rejected():
    with pytest.raises(ValueError, match="one owning floor"):
        validate_records(
            [
                {"Item_ID": "K-1", "Owner": "Oracle"},
                {"Item_ID": "K-1", "Owner": "Morpheus"},
            ]
        )


def test_drive_artifacts_include_neo_six_floors_and_exchange_targets(tmp_path):
    result = write_drive_artifacts(tmp_path)
    manifest = json.loads(
        (tmp_path / "manifest" / "drive_targets.json").read_text(
            encoding="utf-8"
        )
    )

    assert set(result["workbooks"]) == {
        "N. LightSpeed",
        "Trinity",
        "Architect",
        "Morpheous",
        "Oracle",
        "Smith",
        "Merovingian",
        "LS Neo",
        "LS Web",
        "LS GO",
    }
    assert manifest["targets"]["N. LightSpeed"]["folder_id"] == (
        "1lGoz1i8UBbYEt7lCl1esRg7IfF3it24q"
    )
    assert manifest["targets"]["Trinity"]["folder_id"] == (
        "1fivprxdxIahCsWSZXTbVdFfJ5m3FCnf3"
    )
    assert "The Construct" not in manifest["targets"]
    assert (
        tmp_path / "landings" / "N_LIGHTSPEED_README.md"
    ).read_text(encoding="utf-8").startswith("# N. LightSpeed")
