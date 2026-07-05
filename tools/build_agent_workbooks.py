from __future__ import annotations

import argparse
from datetime import UTC, datetime
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, quote_sheetname


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_ROOT = REPO_ROOT / "drive"
GIT_REPOSITORY = "https://github.com/achillesromer-coder/LightSpeed"

TEAL = "156162"
SECONDARY = "235160"
CHARCOAL = "0A0C0B"
OFF_WHITE = "ECE4E2"
CREAM = "F7F3E8"
GOLD = "C9A24A"
GREY = "A7A9AC"
BLUE_GREY = "4C6873"

CLASSIFICATION_COLORS = {
    "Known": "B7FF59",
    "Derived": "3FA65B",
    "Hypothesis": "F2C94C",
    "Conceptual Model": "5A2A4A",
    "Requires Validation": "D66A6A",
}
RELEASE_COLORS = {
    "PUBLIC-CANDIDATE": "B7FF59",
    "Public-filtered": "3FA65B",
    "RI-INTERNAL": "C9A24A",
    "HIGH-SECURITY": "D66A6A",
    "MIXED": "4C6873",
    "CONTROL": "156162",
}

CORE_COLUMNS = [
    "Item_ID",
    "Workstream",
    "Artifact",
    "Input_Source",
    "Local_Path",
    "Drive_URL",
    "Release_Class",
    "Claim_Grade",
    "Status",
    "Completion",
    "Owner",
    "Last_Review_UTC",
    "Next_Action",
    "Evidence_Notes",
    "Cross_Floor_Refs",
    "Notes",
    "Extensions_JSON",
]

NEO_SHEETS = [
    "00 Dashboard",
    "01 Index",
    "02 Floor Registry",
    "03 Projects",
    "04 Quick Knowns",
    "05 Open Tasks",
    "06 Handoffs",
    "07 Approvals",
    "08 Sync Receipts",
    "09 Exceptions",
    "10 Human Review",
    "11 Aesthetic Lock",
    "12 Daily Log",
    "13 Legend",
]
FLOOR_SHEETS = [
    "00 Dashboard",
    "01 Index",
    "02 Mission",
    "03 Projects",
    "04 Knowns",
    "05 Open Tasks",
    "06 Handoffs",
    "07 Evidence",
    "08 Approvals",
    "09 Sync Receipts",
    "10 Human Review",
    "11 Daily Log",
    "12 Legend",
]
EXCHANGE_SHEETS = [
    "00 Dashboard",
    "01 Queue",
    "02 Handoffs",
    "03 Receipts",
    "04 Exceptions",
    "05 Legend",
]

TARGETS = {
    "N. LightSpeed": {
        "folder_id": "1lGoz1i8UBbYEt7lCl1esRg7IfF3it24q",
        "landing": "N_LIGHTSPEED_README.md",
        "workbook": "N_LightSpeed.xlsx",
        "kind": "neo",
        "owner": "Neo",
    },
    "Trinity": {
        "folder_id": "1fivprxdxIahCsWSZXTbVdFfJ5m3FCnf3",
        "landing": "TRINITY_README.md",
        "workbook": "Trinity_LightSpeed.xlsx",
        "kind": "floor",
        "owner": "Trinity",
    },
    "Architect": {
        "folder_id": "1pyLfQaHBYB_b7Uk-8N1x16_4nmc-_rhA",
        "landing": "ARCHITECT_README.md",
        "workbook": "Architect_LightSpeed.xlsx",
        "kind": "floor",
        "owner": "Architect",
    },
    "Morpheous": {
        "folder_id": "1Gpfq5SQ7Ca1SfPR_-1HLM6GAQTNfO8lE",
        "landing": "MORPHEUS_README.md",
        "workbook": "Morpheous_LightSpeed.xlsx",
        "kind": "floor",
        "owner": "Morpheus",
    },
    "Oracle": {
        "folder_id": "1egRJ7sqDAA1nxysvenavx4g2Z_vrjfQx",
        "landing": "ORACLE_README.md",
        "workbook": "Oracle_LightSpeed.xlsx",
        "kind": "floor",
        "owner": "Oracle",
    },
    "Smith": {
        "folder_id": "1VN0Kw_WZtqpY-b3AI5pEXRMFXSB9zppB",
        "landing": "SMITH_README.md",
        "workbook": "Smith_LightSpeed.xlsx",
        "kind": "floor",
        "owner": "Smith",
    },
    "Merovingian": {
        "folder_id": "1I3ObZ-F4Z8T9jAJC3_MKpzdktHAvO2e1",
        "landing": "MEROVINGIAN_README.md",
        "workbook": "Merovingian_LightSpeed.xlsx",
        "kind": "floor",
        "owner": "Merovingian",
    },
    "LS Neo": {
        "folder_id": "1VCDr6P782elSQCd2YFsZI33DNWJ0Z0cN",
        "landing": "LS_NEO_EXCHANGE_README.md",
        "workbook": "LS_Neo_Exchange.xlsx",
        "kind": "exchange",
        "owner": "Neo",
    },
    "LS Web": {
        "folder_id": "1Itg0-L_mHrL4D5DdQf0SobelEAc9LFIJ",
        "landing": "LS_WEB_EXCHANGE_README.md",
        "workbook": "LS_Web_Exchange.xlsx",
        "kind": "exchange",
        "owner": "LS Web",
    },
    "LS GO": {
        "folder_id": "1tW4UcuWtuiJHzhfOife1BFfMGqLGzhkt",
        "landing": "LS_GO_EXCHANGE_README.md",
        "workbook": "LS_GO_Exchange.xlsx",
        "kind": "exchange",
        "owner": "LS GO",
    },
}

MISSIONS = {
    "Neo": "Coordinate bounded cross-floor work, priority, model assignment, and duplicate suppression.",
    "Trinity": "Own the single operator interface, interaction contracts, and human review gates.",
    "Architect": "Own projects, work breakdown, dependencies, sequencing, and publish plans.",
    "Morpheus": "Own knowledge comparison, provenance, contradiction review, and cross-reference.",
    "Oracle": "Own ingestion, evidence classification, knowns, and compact empirical registers.",
    "Smith": "Own bounded automation, execution queues, retries, and result receipts.",
    "Merovingian": "Own diagnostics, telemetry, persistence, retention, and maintenance evidence.",
    "LS Web": "Exchange sanitized Web queue state and publish receipts with Neo.",
    "LS GO": "Exchange private mobile dashboard tasks and receipts with Neo through governed references.",
}


def validate_records(records: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized = [dict(record) for record in records]
    owners: dict[str, str] = {}
    for record in normalized:
        item_id = str(record.get("Item_ID") or "").strip()
        owner = str(record.get("Owner") or "").strip()
        if not item_id:
            continue
        existing = owners.get(item_id)
        if existing and existing != owner:
            raise ValueError(
                f"{item_id} must have one owning floor; found {existing} and {owner}"
            )
        owners[item_id] = owner
    return normalized


def _stamp() -> str:
    return datetime.now(UTC).isoformat(timespec="seconds")


def _seed_records(owner: str, drive_url: str) -> list[dict[str, Any]]:
    mission = MISSIONS.get(owner, MISSIONS["Neo"])
    owner_key = owner.upper().replace(" ", "-")
    records = [
        {
            "Item_ID": f"{owner_key}-MISSION-001",
            "Workstream": "Mission",
            "Artifact": mission,
            "Input_Source": "CORE controls + local canonical source",
            "Local_Path": "desktop/Desktop_Hooks/LightSpeed",
            "Drive_URL": drive_url,
            "Release_Class": "RI-INTERNAL",
            "Claim_Grade": "Known",
            "Status": "active",
            "Completion": 100,
            "Owner": owner,
            "Last_Review_UTC": _stamp(),
            "Next_Action": "Use this workbook as an index, not a document store.",
            "Evidence_Notes": "Mission derived from the approved Z-floor ownership contract.",
            "Cross_Floor_Refs": "N. LightSpeed",
            "Notes": "",
            "Extensions": {"schema": "lightspeed-agent-memory-v1"},
        },
        {
            "Item_ID": f"{owner_key}-SOURCE-001",
            "Workstream": "Evidence",
            "Artifact": "Canonical LightSpeed source repository",
            "Input_Source": GIT_REPOSITORY,
            "Local_Path": "desktop/source-manifest.json",
            "Drive_URL": drive_url,
            "Release_Class": "PUBLIC-CANDIDATE",
            "Claim_Grade": "Known",
            "Status": "active",
            "Completion": 100,
            "Owner": owner,
            "Last_Review_UTC": _stamp(),
            "Next_Action": "Follow source-manifest hashes for build verification.",
            "Evidence_Notes": "Runtime data, model files, secrets, and archives are excluded.",
            "Cross_Floor_Refs": "Neo; Merovingian",
            "Notes": "",
            "Extensions": {"repository": GIT_REPOSITORY},
        },
    ]
    return records


def _new_workbook(sheet_names: list[str]) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in sheet_names:
        workbook.create_sheet(name)
    return workbook


def _style_shell(sheet, *, title: str, subtitle: str, columns: int) -> None:
    last_column = get_column_letter(max(2, columns))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A1"] = title
    sheet["A2"] = subtitle
    sheet["A1"].font = Font(name="Garamond", size=20, bold=True, color="FFFFFF")
    sheet["A2"].font = Font(name="Garamond", size=12, italic=True, color=OFF_WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=TEAL)
    sheet["A2"].fill = PatternFill("solid", fgColor=SECONDARY)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 24
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def _add_navigation(sheet, sheet_names: list[str]) -> None:
    for column, name in enumerate(sheet_names[:12], start=1):
        cell = sheet.cell(3, column, name.split(" ", 1)[-1])
        cell.hyperlink = f"#{quote_sheetname(name)}!A1"
        cell.style = "Hyperlink"
        cell.font = Font(name="Garamond", size=9, bold=True, color=TEAL)
        cell.fill = PatternFill("solid", fgColor=CREAM)
        cell.alignment = Alignment(horizontal="center")


def _style_headers(sheet, headers: list[str]) -> None:
    border = Border(bottom=Side(style="thin", color=GOLD))
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(4, column, header)
        cell.font = Font(name="Garamond", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=SECONDARY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    sheet.row_dimensions[4].height = 32


def _workstream_sheet(
    workstream: str,
    *,
    neo: bool,
) -> str:
    key = str(workstream or "").strip().casefold()
    mappings = {
        "mission": "02 Floor Registry" if neo else "02 Mission",
        "project": "03 Projects",
        "projects": "03 Projects",
        "known": "04 Quick Knowns" if neo else "04 Knowns",
        "knowns": "04 Quick Knowns" if neo else "04 Knowns",
        "task": "05 Open Tasks",
        "tasks": "05 Open Tasks",
        "handoff": "06 Handoffs",
        "handoffs": "06 Handoffs",
        "evidence": "04 Quick Knowns" if neo else "07 Evidence",
        "approval": "07 Approvals" if neo else "08 Approvals",
        "receipt": "08 Sync Receipts" if neo else "09 Sync Receipts",
        "exception": "09 Exceptions" if neo else "10 Human Review",
        "review": "10 Human Review",
        "log": "12 Daily Log" if neo else "11 Daily Log",
    }
    return mappings.get(key, "04 Quick Knowns" if neo else "04 Knowns")


def _record_values(record: dict[str, Any]) -> list[Any]:
    values = []
    for column in CORE_COLUMNS:
        if column == "Extensions_JSON":
            value = record.get("Extensions_JSON")
            if value is None:
                value = json.dumps(
                    record.get("Extensions") or {},
                    sort_keys=True,
                    ensure_ascii=False,
                )
        else:
            value = record.get(column, "")
        values.append(value)
    return values


def _style_data_rows(sheet, *, start: int = 5) -> None:
    for row in sheet.iter_rows(min_row=start):
        for cell in row:
            cell.font = Font(name="Garamond", size=10, color=CHARCOAL)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        claim = str(sheet.cell(row[0].row, 8).value or "")
        release = str(sheet.cell(row[0].row, 7).value or "")
        if claim in CLASSIFICATION_COLORS:
            sheet.cell(row[0].row, 8).fill = PatternFill(
                "solid",
                fgColor=f"FF{CLASSIFICATION_COLORS[claim]}",
            )
        if release in RELEASE_COLORS:
            sheet.cell(row[0].row, 7).fill = PatternFill(
                "solid",
                fgColor=f"FF{RELEASE_COLORS[release]}",
            )
        drive_cell = sheet.cell(row[0].row, 6)
        if isinstance(drive_cell.value, str) and drive_cell.value.startswith("http"):
            drive_cell.hyperlink = drive_cell.value
            drive_cell.style = "Hyperlink"


def _configure_data_sheet(
    sheet,
    *,
    title: str,
    subtitle: str,
    sheet_names: list[str],
    records: list[dict[str, Any]],
) -> None:
    _style_shell(sheet, title=title, subtitle=subtitle, columns=len(CORE_COLUMNS))
    _add_navigation(sheet, sheet_names)
    _style_headers(sheet, CORE_COLUMNS)
    for record in records:
        sheet.append(_record_values(record))
    _style_data_rows(sheet)
    last_row = max(5, sheet.max_row)
    sheet.auto_filter.ref = f"A4:{get_column_letter(len(CORE_COLUMNS))}{last_row}"
    widths = (20, 16, 40, 32, 40, 34, 20, 20, 15, 12, 18, 24, 38, 48, 30, 32, 45)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _configure_index(workbook: Workbook, owner: str, sheet_names: list[str]) -> None:
    sheet = workbook["01 Index"]
    headers = ["Sheet", "Purpose", "Owner", "Open"]
    _style_shell(
        sheet,
        title=f"{owner} / Hyperlinked Index",
        subtitle="Use links instead of duplicating records between sheets or floors.",
        columns=len(headers),
    )
    _add_navigation(sheet, sheet_names)
    _style_headers(sheet, headers)
    for name in sheet_names:
        row = sheet.max_row + 1
        sheet_name = sheet.cell(row, 1, name)
        sheet_name.hyperlink = f"#{quote_sheetname(name)}!A1"
        sheet_name.style = "Hyperlink"
        sheet.cell(row, 2, name.split(" ", 1)[-1])
        sheet.cell(row, 3, owner)
        link = sheet.cell(row, 4, "Open")
        link.hyperlink = f"#{quote_sheetname(name)}!A1"
        link.style = "Hyperlink"
    _style_data_rows(sheet)
    sheet.auto_filter.ref = f"A4:D{sheet.max_row}"
    for index, width in enumerate((28, 48, 20, 12), start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _configure_dashboard(
    workbook: Workbook,
    owner: str,
    sheet_names: list[str],
    *,
    neo: bool,
) -> None:
    sheet = workbook["00 Dashboard"]
    _style_shell(
        sheet,
        title=f"{owner} / LightSpeed Memory",
        subtitle="Compact human review surface; canonical content remains at the linked owner source.",
        columns=8,
    )
    _add_navigation(sheet, sheet_names)
    _style_headers(sheet, ["Metric", "Value", "Authority", "Review"])
    task_sheet = "05 Open Tasks" if "05 Open Tasks" in sheet_names else "01 Queue"
    approval_sheet = (
        "07 Approvals"
        if neo
        else "08 Approvals"
        if "08 Approvals" in sheet_names
        else "02 Handoffs"
    )
    rows = (
        ("Owner", owner, "Single owner", "Weekly"),
        ("Open task rows", f"=MAX(0,COUNTA('{task_sheet}'!A:A)-1)", "Workbook projection", "Friday 19:00"),
        ("Approval rows", f"=MAX(0,COUNTA('{approval_sheet}'!A:A)-1)", "Workbook projection", "Before publish"),
        ("Repository", GIT_REPOSITORY, "Public source", "Per commit"),
        ("Generated UTC", _stamp(), "Build receipt", "This export"),
    )
    for row in rows:
        sheet.append(row)
    _style_data_rows(sheet)
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 70
    sheet.column_dimensions["C"].width = 28
    sheet.column_dimensions["D"].width = 24
    sheet["B8"].hyperlink = GIT_REPOSITORY
    sheet["B8"].style = "Hyperlink"


def _configure_legend(sheet, *, owner: str, sheet_names: list[str]) -> None:
    headers = ["Classification", "Color", "Meaning", "Release Gate"]
    _style_shell(
        sheet,
        title=f"{owner} / Legend",
        subtitle="Claim grade and release class are independent controls.",
        columns=len(headers),
    )
    _add_navigation(sheet, sheet_names)
    _style_headers(sheet, headers)
    meanings = {
        "Known": "Direct evidence or approved operating fact.",
        "Derived": "Reasoned from known evidence with traceable method.",
        "Hypothesis": "Testable proposition awaiting evidence.",
        "Conceptual Model": "Working model; not an empirical claim.",
        "Requires Validation": "Insufficient evidence for operational use.",
    }
    for label, color in CLASSIFICATION_COLORS.items():
        row = sheet.max_row + 1
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, f"#{color}")
        sheet.cell(row, 2).fill = PatternFill("solid", fgColor=f"FF{color}")
        sheet.cell(row, 3, meanings[label])
        sheet.cell(row, 4, "Human review before promotion")
    start = sheet.max_row + 2
    for label, color in RELEASE_COLORS.items():
        row = sheet.max_row + 1
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, f"#{color}")
        sheet.cell(row, 2).fill = PatternFill("solid", fgColor=f"FF{color}")
        sheet.cell(row, 3, "Release classification")
        sheet.cell(row, 4, "Follow the matching publication boundary")
    _style_data_rows(sheet)
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 58
    sheet.column_dimensions["D"].width = 42
    sheet.auto_filter.ref = f"A4:D{sheet.max_row}"
    sheet.row_dimensions[start].height = 8


def _build_agent_workbook(
    path: Path,
    *,
    owner: str,
    records: Iterable[dict[str, Any]],
    neo: bool,
) -> Path:
    records = validate_records(records)
    sheet_names = NEO_SHEETS if neo else FLOOR_SHEETS
    workbook = _new_workbook(sheet_names)
    _configure_dashboard(workbook, owner, sheet_names, neo=neo)
    _configure_index(workbook, owner, sheet_names)

    by_sheet: dict[str, list[dict[str, Any]]] = {
        name: []
        for name in sheet_names
    }
    for record in records:
        requested = str(record.get("Sheet") or "")
        target = requested if requested in by_sheet else _workstream_sheet(
            str(record.get("Workstream") or ""),
            neo=neo,
        )
        by_sheet[target].append(record)

    special = {"00 Dashboard", "01 Index", sheet_names[-1]}
    for name in sheet_names:
        if name in special:
            continue
        _configure_data_sheet(
            workbook[name],
            title=f"{owner} / {name.split(' ', 1)[-1]}",
            subtitle="Owner-indexed rows with source links; optional fields may remain blank.",
            sheet_names=sheet_names,
            records=by_sheet[name],
        )
    _configure_legend(
        workbook[sheet_names[-1]],
        owner=owner,
        sheet_names=sheet_names,
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def build_neo_workbook(path: Path, records: Iterable[dict[str, Any]]) -> Path:
    return _build_agent_workbook(
        Path(path),
        owner="Neo",
        records=records,
        neo=True,
    )


def build_floor_workbook(
    path: Path,
    *,
    floor: str,
    records: Iterable[dict[str, Any]],
) -> Path:
    return _build_agent_workbook(
        Path(path),
        owner=floor,
        records=records,
        neo=False,
    )


def build_exchange_workbook(
    path: Path,
    *,
    owner: str,
    records: Iterable[dict[str, Any]],
) -> Path:
    records = validate_records(records)
    workbook = _new_workbook(EXCHANGE_SHEETS)
    _configure_dashboard(workbook, owner, EXCHANGE_SHEETS, neo=False)
    queue_map = {
        "handoff": "02 Handoffs",
        "receipt": "03 Receipts",
        "exception": "04 Exceptions",
    }
    grouped = {name: [] for name in EXCHANGE_SHEETS}
    for record in records:
        grouped[queue_map.get(str(record.get("Workstream") or "").casefold(), "01 Queue")].append(record)
    for name in EXCHANGE_SHEETS[1:-1]:
        _configure_data_sheet(
            workbook[name],
            title=f"{owner} / {name.split(' ', 1)[-1]}",
            subtitle="Operational exchange projection; private payloads remain in Drive owner files.",
            sheet_names=EXCHANGE_SHEETS,
            records=grouped[name],
        )
    _configure_legend(
        workbook["05 Legend"],
        owner=owner,
        sheet_names=EXCHANGE_SHEETS,
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _landing_text(target_name: str, target: dict[str, str]) -> str:
    owner = target["owner"]
    mission = MISSIONS.get(owner, MISSIONS["Neo"])
    surface = target["kind"]
    return f"""# {target_name}

## Purpose
{mission}

## Authority
- This folder is the approved long-term {surface} persistence target for {owner}.
- The workbook is a compact index and review surface, not an independent truth store.
- Canonical source: {GIT_REPOSITORY}
- Canonical local runtime: `C:\\LightSpeed_Consolidated\\Desktop_Hooks\\LightSpeed`

## Write Contract
- Keep one owning floor per fact or artifact.
- Use cross-floor references instead of duplicating document bodies.
- Preserve `Item_ID`, source, release class, claim grade, status, evidence, and next action.
- Put short-lived work in the workbook queue sheets; promote reviewed results into owner rows.
- Restricted material remains local or in private Drive and never enters public Git or Athene.
- The Construct is Desktop-bound and has no Drive memory workbook.

## Review
- Neo coordinates cross-surface handoffs.
- Human review is required for restricted declassification and public publication.
- Weekly cleanup runs Friday at 19:00 Australia/Brisbane.
"""


def write_drive_artifacts(output_root: Path = DEFAULT_OUTPUT_ROOT) -> dict[str, Any]:
    output_root = Path(output_root)
    landings = output_root / "landings"
    workbooks = output_root / "workbooks"
    manifest_dir = output_root / "manifest"
    landings.mkdir(parents=True, exist_ok=True)
    workbooks.mkdir(parents=True, exist_ok=True)
    manifest_dir.mkdir(parents=True, exist_ok=True)

    generated: dict[str, str] = {}
    manifest_targets: dict[str, dict[str, Any]] = {}
    for name, target in TARGETS.items():
        drive_url = f"https://drive.google.com/drive/folders/{target['folder_id']}"
        landing_path = landings / target["landing"]
        landing_path.write_text(
            _landing_text(name, target),
            encoding="utf-8",
        )
        workbook_path = workbooks / target["workbook"]
        records = _seed_records(target["owner"], drive_url)
        if target["kind"] == "neo":
            build_neo_workbook(workbook_path, records)
        elif target["kind"] == "floor":
            build_floor_workbook(
                workbook_path,
                floor=target["owner"],
                records=records,
            )
        else:
            build_exchange_workbook(
                workbook_path,
                owner=target["owner"],
                records=records,
            )
        generated[name] = str(workbook_path)
        manifest_targets[name] = {
            **target,
            "folder_url": drive_url,
            "landing_path": str(landing_path.relative_to(output_root)),
            "workbook_path": str(workbook_path.relative_to(output_root)),
        }

    manifest = {
        "schema_version": "lightspeed-drive-targets-v1",
        "generated_at": _stamp(),
        "repository": GIT_REPOSITORY,
        "authority": {
            "local_runtime": r"C:\LightSpeed_Consolidated\Desktop_Hooks\LightSpeed",
            "drive": "approved long-term inter-platform persistence",
            "git": "source, schemas, manifests, and public-safe projections",
        },
        "targets": manifest_targets,
    }
    (manifest_dir / "drive_targets.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    return {
        "output_root": str(output_root),
        "workbooks": generated,
        "manifest": str(manifest_dir / "drive_targets.json"),
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build compact LightSpeed Drive landing and workbook artifacts."
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        default=DEFAULT_OUTPUT_ROOT,
    )
    args = parser.parse_args()
    print(
        json.dumps(
            write_drive_artifacts(args.output_root),
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
